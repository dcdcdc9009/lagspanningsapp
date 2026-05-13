"""Excel-export för materiallista."""
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

PURPLE      = '5C2D90'
PURPLE_DARK = '331751'
PURPLE_PALE = 'F5F2FA'
PURPLE_LITE = 'E8E3F5'
LIGHT_GRAY  = 'F1F1F3'
MED_GRAY    = 'C7C7CC'
WHITE       = 'FFFFFF'
DARK        = '19191E'
GRAY        = '6B707A'

_thin = Side(style='thin', color=MED_GRAY)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _cell(ws, row, col, value, bold=False, bg=None, fg=DARK,
          align='left', wrap=False, font_size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, color=fg, size=font_size)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center',
                             wrap_text=wrap)
    c.border = _border
    return c


def skapa_materiallista_excel(projekt, protokoll_lista, enr_lookup):
    """
    enr_lookup: dict artikelnamn → artikelnummer (E-nummer)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materiallista'

    pnr   = projekt.get('projektnummer', '')
    pnamn = projekt.get('projektnamn', '')
    bered = projekt.get('beredare', '')
    datum = datetime.now().strftime('%Y-%m-%d')

    # ── Titel ────────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f'Materiallista – {pnr} – {pnamn}'
    c.font  = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill  = PatternFill('solid', fgColor=PURPLE_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f'Beredare: {bered}   |   Utskrivet: {datum}'
    c.font  = Font(name='Arial', size=9, color=GRAY)
    c.fill  = PatternFill('solid', fgColor=PURPLE_PALE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Aggregera och gruppera ────────────────────────────────────────
    aggregat = {}
    for bp in protokoll_lista:
        for r in bp.get('rader', []):
            key = (r.get('artikel_id'), r.get('artikelnamn', ''))
            if key not in aggregat:
                aggregat[key] = {**r, 'antal': 0.0}
            aggregat[key]['antal'] += r.get('antal', 0)

    # Berika med E-nummer
    for row in aggregat.values():
        if not row.get('artikelnummer'):
            row['artikelnummer'] = enr_lookup.get(row.get('artikelnamn', ''))

    sorterade = sorted(aggregat.values(),
                       key=lambda x: (x.get('kategori') or '', x.get('artikelnamn') or ''))

    grupper = {}
    for r in sorterade:
        grupper.setdefault(r.get('kategori') or 'Övrigt', []).append(r)

    # ── Kolumnrubriker ────────────────────────────────────────────────
    COLS = ['Artikel', 'E-nummer', 'Kategori', 'Enhet', 'Antal']
    row_nr = 4
    for ci, h in enumerate(COLS, start=1):
        _cell(ws, row_nr, ci, h, bold=True, bg=PURPLE, fg=WHITE,
              align='center' if ci >= 4 else 'left')
    ws.row_dimensions[row_nr].height = 18
    row_nr += 1

    # ── Data per kategori ─────────────────────────────────────────────
    for kat, rader in grupper.items():
        # Kategorirubrik
        ws.merge_cells(
            start_row=row_nr, start_column=1,
            end_row=row_nr,   end_column=len(COLS)
        )
        c = ws.cell(row=row_nr, column=1, value=kat)
        c.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=PURPLE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = _border
        ws.row_dimensions[row_nr].height = 16
        row_nr += 1

        for i, r in enumerate(rader):
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            antal = r.get('antal', 0)
            _cell(ws, row_nr, 1, r.get('artikelnamn', ''), bg=bg, wrap=True)
            _cell(ws, row_nr, 2, r.get('artikelnummer') or '–', bg=bg, fg=GRAY, font_size=9)
            _cell(ws, row_nr, 3, r.get('kategori') or '', bg=bg, fg=GRAY, font_size=9)
            _cell(ws, row_nr, 4, r.get('enhet', ''), bg=bg, align='center')
            c5 = _cell(ws, row_nr, 5, antal, bg=bg, align='right')
            c5.number_format = '#,##0.###'
            ws.row_dimensions[row_nr].height = 15
            row_nr += 1

    # ── Kolumnbredder ─────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    # ── Frys rubrikrad ────────────────────────────────────────────────
    ws.freeze_panes = 'A5'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def skapa_konstruktioner_materiallista_excel(konstruktioner, projekt_namn, enr_lookup):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materiallista'

    datum = datetime.now().strftime('%Y-%m-%d')

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f'Konstruktioner – Materiallista{" – " + projekt_namn if projekt_namn else ""}'
    c.font  = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill  = PatternFill('solid', fgColor=PURPLE_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f'Utskrivet: {datum}'
    c.font  = Font(name='Arial', size=9, color=GRAY)
    c.fill  = PatternFill('solid', fgColor=PURPLE_PALE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Aggregera per typ
    typer_ordning = ['Kabelskåp', 'Kabelförläggning', 'Nätstation', 'Övrigt']
    grupper = {}
    for k in konstruktioner:
        grupper.setdefault(k.get('typ', 'Övrigt'), []).append(k)

    sorterade_typer  = [t for t in typer_ordning if t in grupper]
    sorterade_typer += [t for t in grupper if t not in sorterade_typer]

    COLS = ['Artikel', 'E-nummer', 'Enhet', 'Antal', 'Används i']
    row_nr = 4
    for ci, h in enumerate(COLS, start=1):
        _cell(ws, row_nr, ci, h, bold=True, bg=PURPLE, fg=WHITE,
              align='center' if ci >= 3 else 'left')
    ws.row_dimensions[row_nr].height = 18
    row_nr += 1

    for typ in sorterade_typer:
        aggregat = {}
        for k in grupper[typ]:
            for r in k.get('rader', []):
                key = (r.get('artikelnamn') or '').strip()
                if not key:
                    continue
                if key not in aggregat:
                    aggregat[key] = {
                        'artikelnamn': key,
                        'enhet': r.get('enhet', ''),
                        'antal': 0.0,
                        'konstruktioner': [],
                        'artikelnummer': enr_lookup.get(key),
                    }
                aggregat[key]['antal'] += float(r.get('antal', 0) or 0)
                knamn = k.get('namn', '')
                if knamn not in aggregat[key]['konstruktioner']:
                    aggregat[key]['konstruktioner'].append(knamn)

        if not aggregat:
            continue

        ws.merge_cells(
            start_row=row_nr, start_column=1,
            end_row=row_nr,   end_column=len(COLS)
        )
        c = ws.cell(row=row_nr, column=1, value=typ)
        c.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=PURPLE)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = _border
        ws.row_dimensions[row_nr].height = 16
        row_nr += 1

        for i, art in enumerate(sorted(aggregat.values(), key=lambda x: x['artikelnamn'])):
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            anv = ', '.join(art['konstruktioner'][:5])
            if len(art['konstruktioner']) > 5:
                anv += f" +{len(art['konstruktioner'])-5} till"
            _cell(ws, row_nr, 1, art['artikelnamn'], bg=bg, wrap=True)
            _cell(ws, row_nr, 2, art['artikelnummer'] or '–', bg=bg, fg=GRAY, font_size=9)
            _cell(ws, row_nr, 3, art['enhet'], bg=bg, align='center')
            c4 = _cell(ws, row_nr, 4, art['antal'], bg=bg, align='right')
            c4.number_format = '#,##0.###'
            _cell(ws, row_nr, 5, anv, bg=bg, fg=GRAY, font_size=8, wrap=True)
            ws.row_dimensions[row_nr].height = 15
            row_nr += 1

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 35

    ws.freeze_panes = 'A5'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
